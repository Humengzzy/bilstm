from openpyxl import *
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties

class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


if __name__ == "__main__":
    #1=人
    excel_op1 = ExcelOp(file="blah2_mean_auc.xlsx")
    mean_fpr1 = excel_op1.get_col_value(1)
    print(mean_fpr1)
    mean_tpr1 = excel_op1.get_col_value(2)
    print(mean_tpr1)
    #2=小鼠
    excel_op2 = ExcelOp(file="blam2_mean_auc.xlsx")
    mean_fpr2 = excel_op2.get_col_value(1)
    print(mean_fpr2)
    mean_tpr2 = excel_op2.get_col_value(2)
    print(mean_tpr2)
    #3=酿酒酵母
    excel_op3 = ExcelOp(file="blas2_mean_auc.xlsx")
    mean_fpr3 = excel_op3.get_col_value(1)
    print(mean_fpr3)
    mean_tpr3 = excel_op3.get_col_value(2)
    print(mean_tpr3)
    #4=拟南芥
    excel_op4 = ExcelOp(file="blaa2_mean_auc.xlsx")
    mean_fpr4 = excel_op4.get_col_value(1)
    print(mean_fpr4)
    mean_tpr4 = excel_op4.get_col_value(2)
    print(mean_tpr4)


    plt.plot([0, 1], [0, 1], linestyle='--', lw=2, color='r', label='Random', alpha=.8)
    plt.plot(mean_fpr1, mean_tpr1, color='c',label=r'H.Sapiens ROC (AUC=0.925)',lw=2, alpha=.8)
    plt.plot(mean_fpr2, mean_tpr2, color='g', label=r'M.musculus ROC (AUC=0.989)', lw=2, alpha=.8)
    plt.plot(mean_fpr3, mean_tpr3, color='y', label=r'S.cerevisiae ROC (AUC=0.978)', lw=2, alpha=.8)
    plt.plot(mean_fpr4, mean_tpr4, color='m', label=r'A.thaliana ROC (AUC=0.865)', lw=2, alpha=.8)

    plt.xlim([0, 1.05])
    plt.ylim([0, 1.05])
    font_set = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=12)
    plt.xlabel(u'假阳性率',fontproperties=font_set)
    plt.ylabel(u'真阳性率',fontproperties=font_set)
    #plt.title(u'ROC曲线',fontproperties=font_set)
    plt.legend(loc="lower right")
    title = 'ROC2'
    plt.savefig(title+ '.jpg')

